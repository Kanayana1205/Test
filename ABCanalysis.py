# モジュールのインポート
import os, tkinter, tkinter.filedialog, tkinter.messagebox
import openpyxl as px
import pandas as pd
import csv
from openpyxl.utils import column_index_from_string
from openpyxl.styles.alignment import Alignment

# ファイル選択ダイアログの表示
root = tkinter.Tk()
root.withdraw()

# CSVファイル指定   
fTyp = [("","*.csv")]

iDir = os.path.abspath(os.path.dirname(__file__))
tkinter.messagebox.showinfo("ABC分析プログラム","処理ファイルを選択してください！")
file = tkinter.filedialog.askopenfilename(filetypes = fTyp,initialdir = iDir)



#データをdfに読み込み。pandasをpdとして利用。
df = pd.read_csv(file,encoding="cp932",names=("JANcode", "商品名", "分析値","構成比","累計構成比","判定"))

#エクセルファイルを作成・データフレームの書き込み
df.to_excel("ABCanalysis.xlsx",sheet_name="ABCanalysis",index=False)
wb = px.load_workbook("ABCanalysis.xlsx")
ws = wb["ABCanalysis"]

#最終行を取得
max_row = ws.max_row 

#最終行の次の行に合計・合計値（数式）を記入
ws.cell(row = max_row + 1,column= 2).value = "合計"
ws.cell(row = max_row + 1,column= 2).alignment = Alignment(horizontal='center',
                                                           vertical='bottom')
ws.cell(row = max_row + 1,column= 3).value = "=SUM(C{}:C{})".format(2,max_row)




#書式設定を変更　
for row in range(2,max_row + 2):
    ws.cell(row = row,column = 1).number_format = "0"
    ws.cell(row = row,column = 1).alignment = Alignment(horizontal='center',
                                                          vertical='bottom')
    ws.cell(row = row,column = 3).number_format = "#,##0"
    
    #構成比の数式の記入・書式設定（％）
    ws.cell(row = row,column = 4).value = "=C{}/$C${}".format(row,max_row + 1)
    ws.cell(row = row,column = 4).number_format = "0.00%"


#列の幅を変更
ws.column_dimensions["A"].width = 15
ws.column_dimensions["B"].width = 30
ws.column_dimensions["C"].width = 15
ws.column_dimensions["D"].width = 15
ws.column_dimensions["E"].width = 15



wb.save("ABCanalysis.xlsx")


