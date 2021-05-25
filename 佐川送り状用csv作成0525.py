import csv
import datetime
import re
#import xlwings as xw
import openpyxl as px
import mojimoji
import glob
from tkinter import messagebox
import tkinter as tk
root = tk.Tk()
root.withdraw()

towa_tel = "073-482-4421"
towa_postalcode = "642-0034"
towa_address = "和歌山県海南市藤白７５９"
towa_name = "東和産業株式会社"

now = datetime.datetime.now()
filename = '佐川送り状データ_' + now.strftime('%Y%m%d') + '.csv'





#指定フォルダ内のxlsxファイルを全て取得
#files = glob.glob('X:/通箱/★システム管理部/西澤/佐川送り状/*.xls*')
files = glob.glob("C:/Users/TWHN26/Desktop/CSV転記テスト/*.xls*")

#---------------------------------------------------------------------------------------------------------------------
#まず、入力エラー（文字制限）を全ファイルチェックし、該当ファイルをメッセージボックスに表示させる。なければ続行-----------------

error_file = []

for file in files:
    wb = px.load_workbook(file, data_only=True)
    ws = wb.worksheets[0]

    if not ws.cell(row = 5,column = 6).value == "着日" :
        error_file.append(file)

    for i in range(8,13):
        GetValue = ws.cell(row = i,column = 3).value
        if not GetValue is None:
            if len(GetValue) > 16 :
                error_file.append(file)

#エラーとみなしたファイルがある場合、メッセージボックスに表示
if not len(error_file) == 0:
    messagebox.showerror("【入力エラー】 文字数制限オーバーもしくは過去の書式使用",error_file)
    exit()



#-----------------------------------------------------------------------------------------------------------
for file in files:
    wb = px.load_workbook(file, data_only=True)
    ws = wb.worksheets[0]
    ex_postal_code = ws["C7"].value
    #郵便番号に全角が入っていた場合（空白でなければ）半角にする
    if not ex_postal_code is None:
        ex_postal_code = mojimoji.zen_to_han(ex_postal_code)
    postal_code = ex_postal_code
    address_1 = ws["C8"].value
    address_2 = ws["C9"].value
    address_3 = ws["C10"].value
   

    to_name_1 = ws["C11"].value
    to_name_2 = ws["C12"].value
    tel = ws["C13"].value
    title = ws["C14"].value

    #日付の型をｙｙｙｙｍｍｄｄになるように置き換え
    to_date = ws["G5"].value
    if not to_date is None:
        day = (str(to_date.year) + str(format(int(to_date.month),'02')) + str(format(int(to_date.day),'02')))
    else:
        day = ""


    morning = ws["I7"].value
    if morning == "〇":
        #morning = '="01"'
        morning = "01"
        #エクセルでCSVを開くと「1」となるが、メモ帳で確認すると「01」であり、佐川も01で読み込む
       
    #csvへ入力する為のリスト
    data = [tel,postal_code,address_1,address_2,address_3,\
        to_name_1,to_name_2,'','','','',towa_tel,towa_postalcode,towa_address,'',\
        towa_name,'','001',title,'','','','',1,'000','001',day,morning,'',0,'',1,'',0,'','','',0,0,'',1]

  

    with open(filename, 'a',newline="") as f:
        writer = csv.writer(f)
        writer.writerow(data)

 