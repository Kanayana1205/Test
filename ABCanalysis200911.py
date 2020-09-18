# モジュールのインポート
import os, tkinter, tkinter.filedialog, tkinter.messagebox
import openpyxl as px
import xlwings as xw
import pandas as pd
import csv
from openpyxl.utils import column_index_from_string
from openpyxl.styles.alignment import Alignment
from openpyxl.formatting.rule import DataBarRule
import time

# ファイル選択ダイアログの表示
root = tkinter.Tk()
root.withdraw()

# CSVファイル指定   
fTyp = [("","*.csv")]

iDir = os.path.abspath(os.path.dirname(__file__))
tkinter.messagebox.showinfo("ABC分析プログラム","処理ファイルを選択してください！")
file = tkinter.filedialog.askopenfilename(filetypes = fTyp,initialdir = iDir)


#データをdfに読み込み。pandasをpdとして利用。
df = pd.read_csv(file,encoding="cp932",names=("JANcode", "商品名", "分析値","構成比","累計構成比","判定"))
#dfを降順に並び替え
df_s = df.sort_values("分析値", ascending=False)

#エクセルファイルを作成・データフレームの書き込み
df_s.to_excel("ABCanalysis.xlsx",sheet_name="ABCanalysis",index=False)
wb = px.load_workbook("ABCanalysis.xlsx")
ws = wb["ABCanalysis"]

#最終行を取得
max_row = ws.max_row 

#最終行の次の行に合計・合計値（数式）を記入
ws.cell(row = max_row + 1,column= 2).value = "合計"
ws.cell(row = max_row + 1,column= 2).alignment = Alignment(horizontal='center',
                                                           vertical='bottom')
ws.cell(row = max_row + 1,column= 3).value = "=SUM(C{}:C{})".format(2,max_row)



for row in range(2,max_row + 2):
    #書式設定を変更　
    ws.cell(row = row,column = 1).number_format = "0"
    ws.cell(row = row,column = 1).alignment = Alignment(horizontal='center')
    ws.cell(row = row,column = 3).number_format = "#,##0"
    ws.cell(row = row,column = 6).alignment = Alignment(horizontal='center')

    #構成比の数式の記入・書式設定（％）
    ws.cell(row = row,column = 4).value = "=C{}/$C${}".format(row,max_row + 1)
    ws.cell(row = row,column = 4).number_format = "0.00%"
    ws.cell(row = row,column = 5).number_format = "0.00%" #累計構成比の書式設定


#累計構成比を記入
ws.cell(row = 2,column = 5).value = ws.cell(row = 2,column = 4).value
for row in range(3,max_row + 1):
    ws.cell(row = row,column = 5).value = "=E{}+D{}".format(row-1,row)

#列の幅を変更
ws.column_dimensions["A"].width = 15
ws.column_dimensions["B"].width = 30
ws.column_dimensions["C"].width = 15
ws.column_dimensions["D"].width = 15
ws.column_dimensions["E"].width = 15

#データバーの設定
rule = DataBarRule(start_type="percentile", start_value= 0, end_type="percentile", end_value=100,
                  color="FF638EC6", showValue="None", minLength=None, maxLength=None)

ws.conditional_formatting.add(f"D2:D{max_row}",rule)
ws.conditional_formatting.add(f"E2:E{max_row}",rule)



#累計構成比を値として読み込むために、一度保存してからxlwingsで一度開く
wb.save("ABCanalysis.xlsx")
wbxl = xw.Book("ABCanalysis.xlsx")
wsxl = wbxl.sheets["ABCanalysis"]

#ランク定義
Arank = 0.7
Brank = 0.9

#ABCランクの記入・色付け
for row in range(2,max_row + 1):
    if wsxl.cells(row,5).value <= Arank:
        wsxl.cells(row,6).value = "A"
        wsxl.cells(row,6).color = (152, 251, 152)
    elif wsxl.cells(row,5).value >= Arank and wsxl.cells(row,5).value <= Brank :
        wsxl.cells(row,6).value = "B"
        wsxl.cells(row,6).color = (255, 222, 173)
    else:
        wsxl.cells(row,6).value = "C"
        wsxl.cells(row,6).color = (255, 192, 203)

wbxl.save("ABCanalysis.xlsx")





